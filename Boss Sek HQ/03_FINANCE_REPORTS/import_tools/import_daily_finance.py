#!/usr/bin/env python3
import argparse
import csv
import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


MIN_IMPORT_DATE = date(2026, 1, 1)  # 1 มกราคม พ.ศ. 2569

BUSINESSES = {
    "SK Suki",
    "มาจิเมะ",
    "ร้าน เฮงเฮง ซุปเปอร์สโตร์",
    "Evara",
}

DAILY_COLUMNS = {
    "date": 1,
    "business": 2,
    "revenue_cash": 3,
    "revenue_transfer": 4,
    "revenue_card_qr": 5,
    "revenue_delivery_app": 6,
    "cogs_food_cost": 8,
    "staff_cost": 9,
    "rent": 10,
    "utilities": 11,
    "marketing": 12,
    "supplies": 13,
    "other_expense": 14,
    "orders_bills": 20,
    "customers": 21,
    "notes": 22,
    "source_link": 23,
}


def parse_amount(value):
    if value is None:
        return None
    value = str(value).strip().replace(",", "")
    if not value:
        return None
    return float(value)


def parse_date(value):
    value = str(value).strip()
    if not value:
        raise ValueError("date is required")
    return datetime.strptime(value, "%Y-%m-%d").date()


def first_empty_data_row(ws, key_cols=(1, 2), start=2):
    row = start
    while row <= ws.max_row:
        if all(ws.cell(row, col).value in (None, "") for col in key_cols):
            return row
        row += 1
    return ws.max_row + 1


def normalize_row(raw):
    row = {k: (v.strip() if isinstance(v, str) else v) for k, v in raw.items()}
    business = row.get("business", "")
    if business not in BUSINESSES:
        raise ValueError(f"unsupported business: {business!r}")
    parsed = {
        "date": parse_date(row.get("date")),
        "business": business,
        "notes": row.get("notes") or "",
        "source_link": row.get("source_link") or "",
        "confirm_as_initial_balance": (row.get("confirm_as_initial_balance") or "NO").upper(),
    }
    if parsed["date"] < MIN_IMPORT_DATE:
        raise ValueError(f"date before allowed range: {parsed['date']} < {MIN_IMPORT_DATE}")
    if parsed["date"] > date.today():
        raise ValueError(f"future date is not allowed: {parsed['date']} > {date.today()}")
    for key in DAILY_COLUMNS:
        if key in ("date", "business", "notes", "source_link"):
            continue
        parsed[key] = parse_amount(row.get(key))
    parsed["cash_count"] = parse_amount(row.get("cash_count"))
    parsed["inventory_value"] = parse_amount(row.get("inventory_value"))
    return parsed


def is_blank_or_example(raw):
    if not any((v or "").strip() for v in raw.values() if isinstance(v, str)):
        return True
    return "ตัวอย่าง" in (raw.get("notes") or "")


def existing_daily_key(ws, row):
    target_date = row["date"]
    target_business = row["business"]
    target_source = row["source_link"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value in (None, ""):
            continue
        cell_date = ws.cell(r, 1).value
        if hasattr(cell_date, "date"):
            cell_date = cell_date.date()
        if (
            cell_date == target_date
            and ws.cell(r, 2).value == target_business
            and (not target_source or ws.cell(r, 23).value == target_source)
        ):
            return r
    return None


def write_daily(ws, row):
    existing = existing_daily_key(ws, row)
    target_row = existing or first_empty_data_row(ws)
    for key, col in DAILY_COLUMNS.items():
        if key in row and row[key] is not None:
            ws.cell(target_row, col).value = row[key]
    ws.cell(target_row, 1).number_format = "yyyy-mm-dd"
    for col in range(3, 20):
        ws.cell(target_row, col).number_format = "#,##0.00"
    return "updated" if existing else "added"


def write_capital(cap_ws, row):
    if row.get("cash_count") is None and row.get("inventory_value") is None:
        return None
    target_row = first_empty_data_row(cap_ws)
    cap_ws.cell(target_row, 1).value = row["date"]
    cap_ws.cell(target_row, 2).value = row["business"]
    cap_ws.cell(target_row, 6).value = row.get("inventory_value")
    cap_ws.cell(target_row, 7).value = row.get("cash_count")
    cap_ws.cell(target_row, 8).value = "Daily Discord finance import. Confirm before treating as starting capital."
    cap_ws.cell(target_row, 9).value = row.get("source_link")
    cap_ws.cell(target_row, 1).number_format = "yyyy-mm-dd"
    for col in range(3, 8):
        cap_ws.cell(target_row, col).number_format = "#,##0.00"
    return "added"


def maybe_update_settings(settings_ws, row):
    if row.get("confirm_as_initial_balance") != "YES":
        return False
    for r in range(2, settings_ws.max_row + 1):
        if settings_ws.cell(r, 1).value == row["business"]:
            if row.get("cash_count") is not None:
                settings_ws.cell(r, 3).value = row["cash_count"]
            if row.get("inventory_value") is not None:
                settings_ws.cell(r, 4).value = row["inventory_value"]
            note = "Initial balance updated from confirmed daily import."
            existing = settings_ws.cell(r, 7).value or ""
            if note not in existing:
                settings_ws.cell(r, 7).value = (existing + " | " + note).strip(" |")
            return True
    return False


def append_log(log_ws, csv_path, rows_added, rows_updated, capital_rows, settings_updates, status, notes):
    row = log_ws.max_row + 1
    values = {
        1: datetime.now(),
        2: "Multiple" if rows_added + rows_updated > 1 else "",
        3: "Daily CSV import",
        4: csv_path.name,
        5: "See source rows",
        6: rows_added + rows_updated + capital_rows,
        7: str(csv_path),
        8: "Codex",
        9: status,
        10: f"{notes} Daily added={rows_added}, updated={rows_updated}, capital_rows={capital_rows}, settings_updates={settings_updates}.",
    }
    for col, value in values.items():
        log_ws.cell(row, col).value = value


def main():
    parser = argparse.ArgumentParser(description="Import Boss Sek daily finance CSV into dashboard workbook.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--csv", required=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    csv_path = Path(args.csv)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not csv_path.exists():
        raise SystemExit(f"CSV not found: {csv_path}")

    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = [normalize_row(raw) for raw in reader if not is_blank_or_example(raw)]

    if not rows:
        print("No importable rows found.")
        return

    wb = load_workbook(workbook_path)
    daily_ws = wb["Daily_Input"]
    cap_ws = wb["Capital"]
    settings_ws = wb["Settings"]
    log_ws = wb["Import_Log"]

    daily_added = daily_updated = capital_rows = settings_updates = 0
    for row in rows:
        result = write_daily(daily_ws, row)
        if result == "added":
            daily_added += 1
        else:
            daily_updated += 1
        if write_capital(cap_ws, row):
            capital_rows += 1
        if maybe_update_settings(settings_ws, row):
            settings_updates += 1

    append_log(
        log_ws,
        csv_path,
        daily_added,
        daily_updated,
        capital_rows,
        settings_updates,
        "Dry run" if args.dry_run else "Imported",
        "Dry run only." if args.dry_run else "Imported from daily queue.",
    )

    if args.dry_run:
        print(f"DRY RUN: would add {daily_added}, update {daily_updated}, capital {capital_rows}, settings {settings_updates}.")
        return

    backup = workbook_path.with_name(workbook_path.stem + "_backup_before_daily_import.xlsx")
    if not backup.exists():
        shutil.copy2(workbook_path, backup)
    wb.save(workbook_path)
    print(f"Imported {daily_added} added, {daily_updated} updated, {capital_rows} capital rows, {settings_updates} settings updates.")


if __name__ == "__main__":
    main()
